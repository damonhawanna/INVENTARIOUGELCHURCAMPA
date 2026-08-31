# -*- coding: utf-8 -*-
"""
Genera data.js (INVENTARIOUGELCHURCAMPA) a partir del formato de bienes patrimoniales
fmt_inventario_2026-01.xlsx / fmt_inventario_2026-02.xlsx (convertidos desde .xls con LibreOffice).

Uso:  python generar_datos_fmt.py
Salida: data.js (variable global INVENTARIOUGELCHURCAMPA)
"""
import json
import os
import re
import openpyxl

FUENTES = [
    "../datos-crudo/fmt_inventario_2026-01.xlsx",
    "../datos-crudo/fmt_inventario_2026-02.xlsx",
]

KEEP = [
    ("CODIGO_PATRIMONIAL", "cod"),
    ("DENOMINACION_BIEN", "bien"),
    ("TIPO", "tipo"),
    ("NRO_DOC_ADQUISICION", "doc"),
    ("FECHA_ADQUISICION", "fecha"),
    ("VALOR_ADQUISICION", "vadq"),
    ("VALOR_NETO", "vneto"),
    ("NOMBRE_AREA", "area"),
    ("ESTADO_BIEN", "estado"),
    ("CONDICION", "cond"),
    ("MARCA", "marca"),
    ("MODELO", "modelo"),
    ("SERIE", "serie"),
    ("COLOR", "color"),
    ("APELLIDO_PATERNO", "ap"),
    ("APELLIDO_MATERNO", "am"),
    ("NOMBRES", "nom"),
]

OFICINAS = {
    "ABASTECIMIENTO.", "ADMINISTRACION", "ALMACEN",
    "AREA DE GESTION PEDAGOGICA",
    "AREA DE PLANEAMIENTO, PPTO Y MODERNIZACION",
    "ASESORIA JURICA", "CONTABILIDAD.", "CONVIVENCIA ESCOLAR",
    "DIRECCION.", "EDUCACION INICIAL", "EDUCACION PRIMARIA",
    "EDUCACION SECUNDARIA", "ESTADISTICA", "FINANZAS",
    "INFRAESTRUCTURA", "PATRIMONIO", "PLANIFICACION",
    "RACIONALIZACION", "RECURSOS HUMANOS",
    "REDUCCION DE VULNERABILIDAD Y ATENCION DE EMERGENCIAS POR DESASTRE",
    "REMUNERACIONES Y PENSIONES", "TESORERIA",
}

VALORES_LIMPIO = {"SIN MARCA", "SIN MODELO", "SIN SERIE", "SIN MEDIDAS",
                  "SIN CARACTERISTICAS", "NO INDICA", "NO INDIca", "nan",
                  "NAN", "NaN", "SIN INFORMACION"}


def limpiar(v):
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    if s.lower() in {x.lower() for x in VALORES_LIMPIO}:
        return ""
    return s


def deducir_cat(area):
    a = (area or "").upper()
    if not a:
        return ""
    if a in OFICINAS:
        return "SEDE / UGEL"
    if re.match(r"^\s*B", a) and "ESPECIAL" in a:
        return "CEBE"
    if re.match(r"^\s*T", a) or "CETPRO" in a or "TEC." in a or "TÉC." in a:
        return "CETPRO"
    if re.match(r"^\s*I\.?E\.?I\.?", a):
        return "INICIAL"
    if re.match(r"^\s*I\.?E\.?P\.?", a):
        return "PRIMARIA"
    if re.match(r"^\s*I\.?E\.?S\.?", a):
        return "SECUNDARIA"
    if re.match(r"^\s*I\.?E\.?\s", a):
        return ""
    if "PPRONOEI" in a or "PRONOEI" in a:
        return "PRONOEI"
    if re.match(r"^\s*C\.?E\.?B\.?A", a):
        return "CEBA"
    if re.match(r"^\s*C\.?E\.?B\.?E", a):
        return "CEBE"
    return ""


def leer(fuente):
    wb = openpyxl.load_workbook(fuente, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {h: i for i, h in enumerate(header)}
    regs = []
    for r in rows:
        obj = {}
        for col, key in KEEP:
            if col in idx:
                v = limpiar(r[idx[col]])
                if v:
                    obj[key] = v
        area = obj.get("area", "")
        cat = deducir_cat(area)
        if cat:
            obj["cat"] = cat
        if obj.get("bien") or obj.get("cod"):
            regs.append(obj)
    return regs


def main():
    todos = []
    for f in FUENTES:
        absf = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), f))
        print("Leyendo:", absf)
        todos.extend(leer(absf))
    print("Registros totales:", len(todos))

    inst_by_area = {}
    for r in todos:
        a = r.get("area")
        if not a:
            continue
        if a not in inst_by_area:
            inst_by_area[a] = {"nombre": a, "cat": r.get("cat", "")}
    instituciones = [
        {"nombre": v["nombre"], "cat": v["cat"]}
        for v in inst_by_area.values()
    ]
    instituciones.sort(key=lambda x: x["nombre"].lower())
    print("Instituciones unicas:", len(instituciones))

    data = {"registros": todos, "instituciones": instituciones}
    js = "window.INVENTARIOUGELCHURCAMPA = " + json.dumps(
        data, ensure_ascii=False, separators=(",", ":")) + ";\n"

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.js")
    with open(out, "w", encoding="utf-8") as f:
        f.write(js)
    print("Escrito:", out, "(", os.path.getsize(out), "bytes )")


if __name__ == "__main__":
    main()
